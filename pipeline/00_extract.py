"""
00_extract.py — Extracción cruda de los libros de Excel de la tesis.

Lee desde THESIS_DATA_DIR (por defecto ./Data_nosubiralrepo) y escribe a
data/private/ , que está en .gitignore. NADA de lo que produce este script
puede llegar al repositorio.

Resuelve dos problemas estructurales de las hojas originales:
  1. Celdas combinadas: el valor vive sólo en la esquina superior izquierda;
     aquí se propaga a todo el rango (por eso lat/prof aparecían con 86% de
     nulos aparentes en la hoja 1).
  2. La columna "Discriminacion Adicional" (microhábitat / vivos-muertos /
     esteras bacterianas) existe en la hoja borrador y se perdió en la hoja
     filtrada. Aquí se recupera y se reincorpora vía (Título, especie).
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.environ.get("THESIS_DATA_DIR", ROOT / "Data_nosubiralrepo"))
OUT_DIR = ROOT / "data" / "private"

BOOK_BIB = "BD FORAMS AMBTE FILTRACION-filtros (1).xlsx"
BOOK_MSH = "Colección -Clasificacion - Conteo MSH-BC-21 (1).xlsx"


def load_grid(ws) -> list[list]:
    """Devuelve la hoja como matriz con las celdas combinadas ya propagadas."""
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    for rng in ws.merged_cells.ranges:
        r0, r1 = rng.min_row - 1, rng.max_row - 1
        c0, c1 = rng.min_col - 1, rng.max_col - 1
        anchor = grid[r0][c0] if r0 < len(grid) and c0 < len(grid[r0]) else None
        for r in range(r0, min(r1 + 1, len(grid))):
            for c in range(c0, min(c1 + 1, len(grid[r]))):
                grid[r][c] = anchor
    return grid


def clean(v):
    if v is None:
        return None
    s = str(v).replace(" ", " ").strip()
    s = " ".join(s.split())
    return s or None


def block_fill(rows: list[dict], field: str) -> None:
    """Propaga hacia abajo el valor de un campo hasta que aparezca uno nuevo.

    Se usa para 'titulo' y para 'discriminacion', que en el original actúan
    como encabezados de bloque: el valor aplica a las filas que le siguen.
    """
    last = None
    for r in rows:
        if r.get(field):
            last = r[field]
        else:
            r[field] = last


def extract_bibliografia(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Hoja borrador: 5 columnas, conserva 'Discriminacion Adicional' ---
    grid = load_grid(wb["BD FORAMS AMBTE FILTRACION"])
    borrador = []
    for i, row in enumerate(grid[1:], start=2):
        vals = [clean(x) for x in (list(row) + [None] * 5)[:5]]
        if not any(vals):
            continue
        borrador.append(
            {
                "fila_excel": i,
                "titulo": vals[0],
                "especie_raw": vals[1],
                "lat_banda": vals[2],
                "prof_banda": vals[3],
                "discriminacion": vals[4],
            }
        )
    block_fill(borrador, "titulo")
    block_fill(borrador, "lat_banda")
    block_fill(borrador, "prof_banda")
    block_fill(borrador, "discriminacion")
    # '-' es "sin discriminación", no un valor
    for r in borrador:
        if r["discriminacion"] == "-":
            r["discriminacion"] = None

    # --- Hoja maestra: 6 columnas, con tipo de pared ---
    grid = load_grid(wb["Biblio filtrada"])
    maestra = []
    for i, row in enumerate(grid[1:], start=2):
        vals = [clean(x) for x in (list(row) + [None] * 6)[:6]]
        if not any(vals):
            continue
        maestra.append(
            {
                "fila_excel": i,
                "titulo": vals[0],
                "especie_raw": vals[1],
                "pared_raw": vals[2],
                "subtipo_raw": vals[3],
                "lat_banda": vals[4],
                "prof_banda": vals[5],
            }
        )
    block_fill(maestra, "titulo")

    # --- Recuperar 'discriminacion' desde el borrador vía (titulo, especie) ---
    idx: dict[tuple, str] = {}
    for r in borrador:
        if r["titulo"] and r["especie_raw"] and r["discriminacion"]:
            idx.setdefault((r["titulo"][:60].lower(), r["especie_raw"].lower()), r["discriminacion"])
    recuperados = 0
    for r in maestra:
        key = ((r["titulo"] or "")[:60].lower(), (r["especie_raw"] or "").lower())
        if key in idx:
            r["discriminacion"] = idx[key]
            recuperados += 1
        else:
            r["discriminacion"] = None

    # --- Hojas de filtro: se verifica que sean subconjuntos, no se cargan ---
    filtros = {}
    for name in wb.sheetnames:
        if name.startswith(("Lat ", "Prof ")):
            g = load_grid(wb[name])
            filtros[name] = sum(1 for row in g[1:] if any(clean(x) for x in row))

    wb.close()
    return {
        "borrador": borrador,
        "maestra": maestra,
        "filtros": filtros,
        "recuperados_discriminacion": recuperados,
    }


def extract_msh(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Clasificacion: bloque A:C taxonómico + D:H de cálculo ---
    ws = wb["Clasificacion"]
    especies = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        v = [clean(x) for x in (list(row) + [None] * 8)[:8]]
        # fila válida = tiene especie (col C) y total numérico (col E)
        if not v[2] or v[4] is None:
            continue
        try:
            total = float(str(v[4]).replace(",", "."))
        except ValueError:
            continue
        especies.append(
            {
                "fila_excel": i,
                "especie_raw": v[2],
                "pared_raw": v[0],
                "subtipo_raw": v[1],
                "total": total,
            }
        )

    # --- Abundancias: encabezado en 2 niveles, 2 bloques lado a lado ---
    ws = wb["Abundancias"]
    g = load_grid(ws)
    fracciones = ["125", "250", "500", "TOTAL"]
    abundancias = []
    # bloque 1 -> cols B..E (idx 1..4); bloque 2 -> cols H..K (idx 7..10)
    for muestra_col, val_c0 in ((0, 1), (6, 7)):
        muestra = clean(g[0][muestra_col])
        for row in g[2:]:
            etiqueta = clean(row[muestra_col])
            if not etiqueta:
                continue
            vals = {}
            for k, frac in enumerate(fracciones):
                raw = row[val_c0 + k] if val_c0 + k < len(row) else None
                try:
                    vals[frac] = float(str(raw).replace(",", ".")) if raw is not None else None
                except ValueError:
                    vals[frac] = None
            abundancias.append({"muestra": muestra, "variable": etiqueta, "valores": vals})

    wb.close()
    return {"especies": especies, "abundancias": abundancias}


def main() -> int:
    if not DATA_DIR.exists():
        print(f"ERROR: no existe {DATA_DIR}", file=sys.stderr)
        print("Define THESIS_DATA_DIR con la ruta a la carpeta de datos.", file=sys.stderr)
        return 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    bib = extract_bibliografia(DATA_DIR / BOOK_BIB)
    msh = extract_msh(DATA_DIR / BOOK_MSH)

    (OUT_DIR / "bibliografia_raw.json").write_text(
        json.dumps(bib, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    (OUT_DIR / "msh_raw.json").write_text(
        json.dumps(msh, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    print("EXTRACCIÓN COMPLETA")
    print(f"  Hoja borrador      : {len(bib['borrador'])} registros, "
          f"{len(set(r['titulo'] for r in bib['borrador'] if r['titulo']))} estudios")
    print(f"  Hoja maestra       : {len(bib['maestra'])} registros, "
          f"{len(set(r['titulo'] for r in bib['maestra'] if r['titulo']))} estudios")
    print(f"  Discriminación recuperada en {bib['recuperados_discriminacion']} registros")
    print(f"  Hojas de filtro    : {bib['filtros']}")
    print(f"  MSH-BC-21          : {len(msh['especies'])} especies, "
          f"{len(msh['abundancias'])} filas de abundancia")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
