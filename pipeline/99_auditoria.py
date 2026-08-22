"""
99_auditoria.py — Auditoría independiente del pipeline.

No confía en las salidas intermedias: vuelve a leer los Excel originales y
verifica, etapa por etapa, que nada se perdió, se duplicó ni se alteró sin
quedar registrado en el log de correcciones.

Cada comprobación imprime OK o FALLA. Si algo falla, el script termina con
código 1. Ejecutar después de cualquier cambio en el pipeline.
"""

from __future__ import annotations

import json
import re
import math
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import corrections as C  # noqa: E402
import taxonomy as T  # noqa: E402
import tipologia as TIP  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.environ.get("THESIS_DATA_DIR", ROOT / "Data_nosubiralrepo"))
PRIV = ROOT / "data" / "private"
DERIV = ROOT / "data" / "derived"

FALLAS: list[str] = []
N_OK = 0


def check(nombre: str, cond: bool, detalle: str = "") -> None:
    global N_OK
    if cond:
        N_OK += 1
        print(f"  OK    {nombre}" + (f"  ({detalle})" if detalle else ""))
    else:
        FALLAS.append(nombre)
        print(f"  FALLA {nombre}  -> {detalle}")


def load(p: Path):
    return json.loads(p.read_text(encoding="utf-8"))


def main() -> int:
    print("=" * 74)
    print("AUDITORÍA DEL PIPELINE")
    print("=" * 74)

    # ---------------------------------------------------------------------
    print("\n[1] LECTURA INDEPENDIENTE DE LOS EXCEL ORIGINALES")
    wb = openpyxl.load_workbook(
        DATA_DIR / "BD FORAMS AMBTE FILTRACION-filtros (1).xlsx", data_only=True)

    # conteo crudo de la hoja maestra, sin ninguna lógica del pipeline
    ws = wb["Biblio filtrada"]
    filas_con_especie = 0
    lat_vals, prof_vals, pared_vals = Counter(), Counter(), Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and str(row[1]).strip():
            filas_con_especie += 1
            lat_vals[str(row[4]).strip() if row[4] else None] += 1
            prof_vals[str(row[5]).strip() if row[5] else None] += 1
            pared_vals[str(row[2]).strip() if row[2] else None] += 1
    check("Hoja maestra tiene 293 filas con especie", filas_con_especie == 293,
          f"contadas {filas_con_especie}")
    check("Ninguna fila de la maestra sin latitud", None not in lat_vals,
          f"nulos={lat_vals.get(None, 0)}")
    check("Ninguna fila de la maestra sin profundidad", None not in prof_vals,
          f"nulos={prof_vals.get(None, 0)}")
    check("Tipo de pared original: 274 calcáreo / 19 aglutinado",
          pared_vals.get("Calcareo") == 274 and pared_vals.get("Aglutinado") == 19,
          f"{dict(pared_vals)}")

    # las 7 hojas de filtro deben ser subconjuntos exactos de la maestra
    total_lat = total_prof = 0
    for name in wb.sheetnames:
        if not name.startswith(("Lat ", "Prof ")):
            continue
        n = sum(1 for r in wb[name].iter_rows(min_row=2, values_only=True)
                if r[1] and str(r[1]).strip())
        if name.startswith("Lat "):
            total_lat += n
        else:
            total_prof += n
    check("Hojas de filtro por latitud suman 293", total_lat == 293, f"suman {total_lat}")
    check("Hojas de filtro por profundidad suman 293", total_prof == 293, f"suman {total_prof}")

    wb2 = openpyxl.load_workbook(
        DATA_DIR / "Colección -Clasificacion - Conteo MSH-BC-21 (1).xlsx", data_only=True)
    ws = wb2["Clasificacion"]
    msh_raw = []
    for row in ws.iter_rows(min_row=2, max_row=53, values_only=True):
        if row[2] and row[4] is not None:
            msh_raw.append((str(row[2]).strip(), str(row[0]).strip(), float(row[4])))
    check("MSH-BC-21 tiene 52 especies", len(msh_raw) == 52, f"contadas {len(msh_raw)}")
    tot_raw = sum(x[2] for x in msh_raw)
    check("Total ponderado MSH = 1214,125", abs(tot_raw - 1214.125) < 1e-6, f"{tot_raw}")
    check("MSH original: 45 calcáreas / 7 aglutinadas (coincide con la tesis, p.31)",
          Counter(x[1] for x in msh_raw).get("Calcareo") == 45,
          f"{dict(Counter(x[1] for x in msh_raw))}")

    # ---------------------------------------------------------------------
    print("\n[2] CONSERVACIÓN DE REGISTROS A TRAVÉS DEL PIPELINE")
    raw = load(PRIV / "bibliografia_raw.json")
    clean = load(PRIV / "bibliografia_clean.json")
    check("Extracción conserva las 293 filas", len(raw["maestra"]) == 293,
          f"{len(raw['maestra'])}")

    excluidos = C.EXCLUIR_PLANCTONICOS | C.EXCLUIR_PLACEHOLDER
    n_excl = sum(1 for r in raw["maestra"]
                 if (r.get("especie_raw") or "").strip() in excluidos)
    log_c = load(DERIV / "correcciones.json")["correcciones"]
    n_dup = sum(1 for c in log_c if c["tipo"] == "duplicado")
    curada = [r for r in clean if not r.get("recuperado")]
    recuperados = [r for r in clean if r.get("recuperado")]

    check("293 - excluidos - duplicados = base curada",
          len(raw["maestra"]) - n_excl - n_dup == len(curada),
          f"293 - {n_excl} - {n_dup} = {len(raw['maestra']) - n_excl - n_dup}, "
          f"curada={len(curada)}")
    check("Los recuperados salen del borrador, no de la hoja maestra",
          all(r["estudio"] not in {m["titulo"] for m in raw["maestra"]}
              for r in recuperados),
          f"{len(recuperados)} recuperados")
    check("Todo recuperado lleva la pared derivada y marcada como tal",
          all(r.get("pared_derivada") for r in recuperados), "")

    # ninguna fila cruda desaparece sin estar excluida o deduplicada
    crudos = Counter((r["titulo"], r["especie_raw"]) for r in raw["maestra"]
                     if (r.get("especie_raw") or "").strip() not in excluidos)
    limpios = Counter((r["estudio"], r["taxon_original"]) for r in curada)
    perdidos = [k for k in crudos if k not in limpios]
    check("Ningún par (estudio, taxón) desaparece por completo", not perdidos,
          f"perdidos={perdidos[:3]}")
    check("La reducción de filas se explica sólo por duplicados",
          sum(crudos.values()) - sum(limpios.values()) == n_dup,
          f"diferencia={sum(crudos.values()) - sum(limpios.values())}, duplicados={n_dup}")

    # tras deduplicar no debe quedar ninguna fila idéntica
    resid = [k for k, v in Counter(
        tuple(r[c] for c in C.CLAVE_DUPLICADO) for r in clean).items() if v > 1]
    check("No quedan duplicados exactos", not resid, f"{len(resid)} residuales")

    # las variantes legítimas (mismo taxón, distinta banda) deben sobrevivir
    variantes = sum(
        len({(r["lat_banda"], r["prof_banda"]) for r in clean
             if r["estudio"] == e and r["taxon"] == t}) - 1
        for e, t in {(r["estudio"], r["taxon"]) for r in clean})
    check("Se conservan las variantes por banda (mismo taxón, distinto estrato)",
          variantes > 0, f"conservadas {variantes}")

    # ---------------------------------------------------------------------
    print("\n[3] TODA DIFERENCIA ESTÁ DOCUMENTADA")
    log = load(DERIV / "correcciones.json")["correcciones"]
    documentados = {c["desde"] for c in log}
    cambios = {r["taxon_original"] for r in clean if r["taxon"] != r["taxon_original"]}
    check("Todo cambio de nombre está en el log", cambios <= documentados,
          f"sin documentar: {sorted(cambios - documentados)}")

    pared_cambiada = set()
    idx_raw = {(r["titulo"], r["especie_raw"]): r for r in raw["maestra"]}
    for r in clean:
        orig = idx_raw.get((r["estudio"], r["taxon_original"]))
        if orig and T.canon_wall(orig["pared_raw"]) != r["pared"]:
            pared_cambiada.add(r["taxon_original"])
    check("Todo cambio de pared está en el log", pared_cambiada <= documentados,
          f"sin documentar: {sorted(pared_cambiada - documentados)}")
    check("Los excluidos están declarados en el log",
          all(e in documentados for e in excluidos),
          f"faltan: {sorted(e for e in excluidos if e not in documentados)}")

    # ---------------------------------------------------------------------
    print("\n[4] ARITMÉTICA RECALCULADA DESDE CERO")
    msh = load(DERIV / "msh_bc21.json")
    cnt = [e["conteo"] for e in msh["especies"]]
    tot = sum(cnt)
    H = -sum((c / tot) * math.log(c / tot) for c in cnt if c > 0)
    check("Total MSH coincide con el Excel", abs(tot - tot_raw) < 1e-6, f"{tot}")
    check("Shannon H' recalculado", abs(H - msh["indices"]["shannon_H"]) < 1e-4,
          f"auditoría={H:.4f} publicado={msh['indices']['shannon_H']}")
    check("Shannon coincide con el valor de la hoja original (3,4327)",
          abs(H - 3.432749200140165) < 1e-3, f"{H:.4f}")
    J = H / math.log(len(cnt))
    check("Equidad J' recalculada", abs(J - msh["indices"]["equidad_J"]) < 1e-4, f"{J:.4f}")
    D = sum((c / tot) ** 2 for c in cnt)
    check("Simpson D recalculado", abs(D - msh["indices"]["simpson_D"]) < 1e-4, f"{D:.4f}")
    check("Abundancias relativas suman 100%",
          abs(sum(e["abundancia_rel"] for e in msh["especies"]) - 100) < 0.05,
          f"{sum(e['abundancia_rel'] for e in msh['especies']):.3f}%")
    check("Porcentajes de pared suman 100%",
          abs(sum(msh["pared"].values()) - 100) < 0.05, f"{sum(msh['pared'].values())}")
    check("Subtipos suman 100%",
          abs(sum(msh["subtipo"].values()) - 100) < 0.05, f"{sum(msh['subtipo'].values())}")
    check("Calcáreo = hialino + porcelanáceo + monocristalino",
          abs(msh["pared"]["Calcareo"] - (msh["subtipo"]["Hialino"]
              + msh["subtipo"]["Porcelanaceo"] + msh["subtipo"]["Monocristalino"])) < 0.05,
          f"{msh['pared']['Calcareo']}")

    # efecto declarado de la corrección de Ammodiscus
    amod = next((e for e in msh["especies"]
                 if e["genero"].lower() == "ammodiscus"), None)
    check("Ammodiscus quedó como aglutinado", amod and amod["pared"] == "Aglutinado",
          f"{amod['pared'] if amod else 'no encontrado'}")
    fba_sin = 100 * sum(e["conteo"] for e in msh["especies"]
                        if e["pared"] == "Aglutinado"
                        and e["genero"].lower() != "ammodiscus") / tot
    check("Sin la corrección el FBA sería 11,2% (valor original)",
          abs(fba_sin - 11.2) < 0.1, f"{fba_sin:.2f}%")

    # abundancias: ind/g = extraídos / peso picado
    ab = msh["abundancias"]
    errs = []
    for m, v in ab.items():
        for frac in ("125", "250", "500", "TOTAL"):
            ext = v["Forams B extraidos"][frac]
            peso = v["Peso picado (g)"][frac]
            pub = v["Abundancia F.B"][frac]
            if ext and peso and pub and abs(ext / peso - pub) > 1.0:
                errs.append(f"{m}/{frac}: {ext}/{peso}={ext/peso:.1f} vs {pub}")
    check("Abundancias = extraídos / peso picado (ind/g)", not errs, "; ".join(errs))

    # ---------------------------------------------------------------------
    print("\n[5] COHERENCIA ENTRE LOS DATASETS PÚBLICOS")
    mat = load(DERIV / "matriz_lat_prof.json")
    tax = load(DERIV / "taxones_global.json")
    est = load(DERIV / "estudios.json")
    par = load(DERIV / "pared_por_banda.json")
    sol = load(DERIV / "solape.json")

    n_ns = mat["registros_no_filtracion"]
    check("Matriz curada + no-filtración = base curada",
          sum(c["registros"] for c in mat["celdas"]) + n_ns == len(curada),
          f"{sum(c['registros'] for c in mat['celdas'])} + {n_ns} vs {len(curada)}")
    check("Matriz ampliada = curada + recuperados",
          sum(c["registros"] for c in mat["celdas_ampliada"])
          == sum(c["registros"] for c in mat["celdas"]) + len(recuperados),
          f"{sum(c['registros'] for c in mat['celdas_ampliada'])}")
    check("La ampliación nunca reduce ninguna celda",
          all(a["registros"] >= b["registros"]
              for a, b in zip(mat["celdas_ampliada"], mat["celdas"])), "")
    check("La celda 0-15 / <150 m sigue vacía incluso ampliando la base",
          next(c for c in mat["celdas_ampliada"]
               if c["lat"] == "0-15" and c["prof"] == "< 150 m")["registros"] == 0, "")
    check("Los estudios no-filtración están marcados en estudios.json",
          sum(e["n_registros"] for e in est if not e["es_filtracion"]) == n_ns,
          f"{sum(e['n_registros'] for e in est if not e['es_filtracion'])} vs {n_ns}")

    # el sitio de la tesis debe apuntar a una celda REAL de la matriz: la banda
    # y la coordenada son campos distintos y no deben pisarse entre sí
    st = mat["sitio_tesis"]
    check("El sitio de la tesis conserva banda y coordenada por separado",
          isinstance(st.get("banda_lat"), str) and isinstance(st.get("lat"), (int, float)),
          f"banda_lat={st.get('banda_lat')!r} lat={st.get('lat')!r}")
    celda_sitio = next((c for c in mat["celdas"]
                        if c["lat"] == st.get("banda_lat")
                        and c["prof"] == st.get("banda_prof")), None)
    check("La banda del sitio corresponde a una celda existente", celda_sitio is not None,
          f"{st.get('banda_lat')} / {st.get('banda_prof')}")
    check("La celda del sitio de la tesis está vacía (es el argumento central)",
          celda_sitio is not None and celda_sitio["registros"] == 0,
          f"registros={celda_sitio['registros'] if celda_sitio else '?'}")
    check("La coordenada del sitio cae en su banda latitudinal declarada",
          celda_sitio is not None and 0 <= st["lat"] <= 15,
          f"lat={st.get('lat')}")
    check("taxones_global suma la base curada de filtraciones",
          sum(t["registros"] for t in tax) == len(curada) - n_ns,
          f"{sum(t['registros'] for t in tax)} vs {len(curada) - n_ns}")
    check("taxones_global ampliada suma la base ampliada",
          sum(t["registros_ampliada"] for t in tax) == len(clean),
          f"{sum(t['registros_ampliada'] for t in tax)} vs {len(clean)}")
    check("estudios.json suma todos los registros limpios",
          sum(e["n_registros"] for e in est) == len(clean),
          f"{sum(e['n_registros'] for e in est)}")
    check("Nº de taxones únicos coherente",
          len(tax) == len({r["taxon"] for r in clean}), f"{len(tax)}")
    check("Nº de estudios coherente",
          len(est) == len({r["estudio"] for r in clean}), f"{len(est)}")
    for eje, valores in (("latitud", 4), ("profundidad", 3)):
        sub = [p for p in par if p["eje"] == eje]
        check(f"pared_por_banda cubre las bandas de {eje}", len(sub) == valores,
              f"{len(sub)}")
        check(f"n de {eje} suma la base curada de filtraciones",
              sum(p["n"] for p in sub) == len(curada) - n_ns,
              f"{sum(p['n'] for p in sub)} vs {len(curada) - n_ns}")
    malos = [p for p in par if p["n"] and abs((p["calcareo"] or 0) + (p["aglutinado"] or 0) - 100) > 0.15]
    check("Porcentajes de pared suman 100 en cada banda", not malos,
          f"{[(p['banda'], p['calcareo'], p['aglutinado']) for p in malos]}")

    # solape recalculado independientemente
    # el solape se publica contra la base CURADA de filtraciones: hay que
    # reproducir esa misma base aquí, no el conjunto completo de registros
    xr = load(PRIV / "estudios_crossref.json")
    lim = {x["titulo_original"]: x["titulo_limpio"] for x in xr}
    titulos_seep = {e["titulo"] for e in est if e["es_filtracion"]}
    seep_curada = [r for r in curada
                   if lim.get(r["estudio"]) in titulos_seep]
    check("La base curada de filtraciones tiene el tamaño esperado",
          len(seep_curada) == len(curada) - n_ns,
          f"{len(seep_curada)} vs {len(curada) - n_ns}")

    gl_bin = {T.binomen(r["taxon"]) for r in seep_curada}
    ms_bin = {T.binomen(e["taxon"]) for e in msh["especies"]}
    check("Solape de especies recalculado",
          len(gl_bin & ms_bin) == sol["especies"]["n_compartidas"],
          f"auditoría={len(gl_bin & ms_bin)} publicado={sol['especies']['n_compartidas']}")
    gl_gen = {r["genero"].lower() for r in seep_curada}
    ms_gen = {e["genero"].lower() for e in msh["especies"]}
    check("Solape de géneros recalculado",
          len(gl_gen & ms_gen) == sol["generos"]["n_compartidos"],
          f"auditoría={len(gl_gen & ms_gen)} publicado={sol['generos']['n_compartidos']}")

    # ---------------------------------------------------------------------
    print("\n[6] INTEGRIDAD DE CAMPOS QUE ALIMENTAN GRÁFICOS")
    sin_pared = [t["taxon"] for t in tax if t["pared"] not in ("Calcareo", "Aglutinado")]
    check("Todo taxón global tiene pared válida", not sin_pared, f"{sin_pared[:5]}")
    sin_pared_msh = [e["taxon"] for e in msh["especies"]
                     if e["pared"] not in ("Calcareo", "Aglutinado")]
    check("Toda especie de MSH tiene pared válida", not sin_pared_msh, f"{sin_pared_msh[:5]}")
    sub_malo = [e["taxon"] for e in msh["especies"]
                if e["pared"] == "Calcareo" and e["subtipo"] not in
                ("Hialino", "Porcelanaceo", "Aragonito", "Monocristalino")]
    check("Todo calcáreo de MSH tiene subtipo válido", not sub_malo, f"{sub_malo[:5]}")
    check("Ningún aglutinado tiene subtipo",
          not [t for t in tax if t["pared"] == "Aglutinado" and t["subtipo"]], "")

    geo = [e for e in est if e["lat"] is not None]
    check("Sólo un estudio sin georreferenciar (McCorkle, multi-sitio y no-seep)",
          len(est) - len(geo) == 1, f"{len(est) - len(geo)} sin coordenada")

    # Comprobación cruzada fuerte: la coordenada de cada estudio debe caer
    # dentro de alguna de las bandas latitudinales que sus propios registros
    # declaran. Detecta a la vez coordenadas mal transcritas y bandas mal
    # asignadas en la hoja original.
    bandas = [((0, 15), "0-15"), ((15, 30), "15-30"),
              ((30, 60), "30-60"), ((60, 90), "60-90")]
    xref = load(PRIV / "estudios_crossref.json")
    limpio_de = {x["titulo_original"]: x["titulo_limpio"] for x in xref}
    por_titulo_limpio = defaultdict(set)
    for r in clean:
        por_titulo_limpio[limpio_de.get(r["estudio"])].add(r["lat_banda"])

    incoherentes = []
    for e in geo:
        decl = por_titulo_limpio.get(e["titulo"])
        if not decl:
            continue
        # un estudio multi-sitio declara varias bandas; basta que la
        # coordenada representativa caiga en alguna de ellas
        banda = next((b for (lo, hi), b in bandas if lo <= abs(e["lat"]) <= hi), None)
        if banda and banda not in decl:
            incoherentes.append(
                f"{e['titulo'][:34]}: {e['lat']:.2f}° = banda {banda}, "
                f"pero sus registros dicen {sorted(decl)}")
    check("La coordenada de cada estudio cae en su banda latitudinal declarada",
          not incoherentes, " | ".join(incoherentes[:3]))

    multi = [e for e in est if e.get("sitios")]
    malos = [e["titulo"][:40] for e in multi for s in e["sitios"]
             if not (-90 <= s["lat"] <= 90 and -180 <= s["lon"] <= 180)]
    check("Sitios secundarios con coordenadas válidas", not malos, f"{malos}")
    fuera = [e["titulo"][:40] for e in geo
             if not (-90 <= e["lat"] <= 90 and -180 <= e["lon"] <= 180)]
    check("Coordenadas dentro de rango válido", not fuera, f"{fuera}")
    sin_conf = [e["titulo"][:40] for e in est if not e.get("confianza")]
    check("Toda entrada declara nivel de confianza", not sin_conf, f"{sin_conf}")
    check("Los no georreferenciados están marcados 'nula'",
          all(e["confianza"] == "nula" for e in est if e["lat"] is None), "")
    falso = [e for e in est if e["doi"] == "10.1021/acsestwater.3c00740.s001"]
    check("El DOI falso (parafinas cloradas) está anulado", not falso, "")
    check("Todas las referencias verificadas",
          all(e["referencia_verificada"] for e in est),
          f"{sum(1 for e in est if not e['referencia_verificada'])} sin verificar")
    check("Las referencias resueltas a mano declaran de dónde salen",
          all(e.get("referencia_fuente") for e in est if not e.get("doi")),
          "")

    # etiquetas listas para mostrar
    gen_min = [t["taxon"] for t in tax if t["genero"] and t["genero"][0].islower()]
    check("Géneros capitalizados en taxones_global", not gen_min, f"{gen_min[:5]}")
    gen_min = [e["taxon"] for e in msh["especies"]
               if e["genero"] and e["genero"][0].islower()]
    check("Géneros capitalizados en msh_bc21", not gen_min, f"{gen_min[:5]}")
    check("Géneros capitalizados en solape",
          all(g[0].isupper() for g in sol["generos"]["compartidos"]), "")
    sin_rango = [t["taxon"] for t in tax if t.get("rango") not in ("especie", "genero")]
    check("Todo taxón declara su rango", not sin_rango, f"{sin_rango[:5]}")
    n_gen = sum(1 for t in tax if t["rango"] == "genero")
    check("Las entradas de nomenclatura abierta están marcadas",
          0 < n_gen < len(tax) * 0.25,
          f"{n_gen} de {len(tax)} son entradas de género")
    check("El ranking global va por nº de estudios",
          all(tax[i]["n_estudios"] >= tax[i + 1]["n_estudios"] for i in range(len(tax) - 1)),
          "")

    micro = {r for t in tax for r in t["microhabitats"]}
    check("Microhábitat usa vocabulario controlado",
          micro <= {v[0] for v in C.DISCRIMINACION_VOCAB.values()}, f"{micro}")

    car = load(DERIV / "caribe_referencia.json")
    check("Referencia Caribe: 5 localidades + MSH-BC-21", len(car) == 6, f"{len(car)}")
    vacias = [c["nombre"] for c in car if not c["taxones"]]
    check("Toda localidad caribeña tiene taxones", not vacias, f"{vacias}")

    # ---------------------------------------------------------------------
    print("\n[7] CONFIDENCIALIDAD")
    # La regla es no publicar los ARCHIVOS, no dejar de citarlos: mencionar la
    # tesis y sus anexos es buena práctica y el autor lo autorizó de forma
    # expresa. Lo que no debe aparecer son rutas del sistema de archivos ni
    # enlaces de descarga, que sí permitirían llegar al documento.
    RUTAS = ("Data_nosubiralrepo/", "Data_nosubiralrepo\\", "C:\\Users", "/Users/",
             "file://", "OneDrive")
    fuga = []
    for f in DERIV.glob("*.json"):
        txt = f.read_text(encoding="utf-8")
        fuga += [f"{f.name} contiene '{m}'" for m in RUTAS if m in txt]
    check("Ningún dataset público expone rutas del sistema de archivos",
          not fuga, f"{fuga}")

    informe = ROOT / "Informe_curacion_datos.pdf"
    if informe.exists():
        from pypdf import PdfReader
        txt = " ".join((pg.extract_text() or "")
                       for pg in PdfReader(str(informe)).pages)
        check("El informe en PDF no expone rutas del sistema de archivos",
              not [m for m in RUTAS if m in txt],
              f"{[m for m in RUTAS if m in txt]}")
        check("El informe cita la tesis como fuente",
              "Mendoza Rivero" in txt, "")

    # El CLI de Vercel NO lee .gitignore: su lista de exclusiones sale sólo de
    # .vercelignore (o .nowignore) más unos pocos valores por defecto. Mientras
    # ese archivo no existió, `vercel deploy` subió el directorio entero —los
    # dos Excel inéditos, los 47 PDF y todo data/private/— a los servidores de
    # Vercel. No estaban servidos en ninguna URL, pero habían salido del
    # equipo. Se detectó consultando la API de despliegues, no razonando sobre
    # el CLI. Esta comprobación existe para que no vuelva a pasar en silencio.
    # Cada morfología asignada debe poder verse en el CUERPO de su artículo.
    # Existe porque dos asignaciones eran inferencias disfrazadas de evidencia:
    # E24 salía como «pingo de hidratos de Storfjordrenna» y su artículo no dice
    # «pingo» ni una vez —estudia los pockmarks de Vestnesa, a 330 km—, y E03
    # salía como «pockmark» cuando esa palabra sólo aparecía en su propia
    # bibliografía, citando a otro trabajo.
    TERMINO = {
        "pockmark": r"pockmark",
        "volcan_lodo": r"mud[\s-]?volcano",
        "monticulo_hidratos": r"\bmound",
        "pingo_hidratos": r"\bpingo",
        "diapiro": r"diapir",
        "escarpe": r"escarpment|\bscarp",
        "tapete_bacteriano": r"bacterial mat|microbial mat|Beggiatoa",
        "banco_bivalvos": r"clam bed|mussel bed|bivalve|vesicomyid|Calyptogena",
        "respiradero_hidrotermal": r"\bvent\b|hydrotherm",
        "campo_filtracion": r"seep(?:age)? (?:field|site|area)",
    }
    ev_pdf = PRIV / "pdfs_evidencia.json"
    if ev_pdf.exists() and DATA_DIR.exists():
        import re as _re
        from pypdf import PdfReader as _R
        arch = {x["estudio_id"]: x["archivo"] for x in load(ev_pdf)
                if x.get("estudio_id")}
        sin_apoyo = []
        for e in load(DERIV / "estudios.json"):
            m, eid = e.get("morfologia"), e["id"]
            if not m or m not in TERMINO or eid not in arch:
                continue
            p = DATA_DIR / arch[eid]
            if not p.exists():
                continue
            try:
                txt = " ".join((pg.extract_text() or "") for pg in _R(str(p)).pages)
            except Exception:  # noqa: BLE001
                continue
            # se corta la bibliografía: ahí los títulos citados nombran
            # morfologías de OTROS trabajos
            corte = txt.lower().rfind("references")
            cuerpo = txt[:corte] if corte > len(txt) * 0.5 else txt
            if not _re.search(TERMINO[m], cuerpo, _re.I) \
                    and eid not in TIP.JUSTIFICADA_POR_LOCALIDAD:
                sin_apoyo.append(f"{eid}:{m}")
        check("Toda morfología asignada aparece en el cuerpo de su artículo",
              not sin_apoyo, f"{sin_apoyo}")
        # Las excepciones son legítimas, pero no pueden acumularse en silencio.
        check("Las morfologías justificadas sólo por la localidad son pocas y "
              "están declaradas",
              len(TIP.JUSTIFICADA_POR_LOCALIDAD) <= 3,
              f"{sorted(TIP.JUSTIFICADA_POR_LOCALIDAD)}")

    # Lo extraído de las tablas se valida por RANGO FÍSICO: una cifra fuera de
    # rango significa que se leyó la columna equivocada, y eso hay que
    # detectarlo aquí y no en el dashboard.
    tab = PRIV / "tablas_pdf.json"
    if tab.exists():
        tt = load(tab)
        d13 = [v["valor"] for v in tt["valores"] if v["variable"] == "d13C"]
        ab = [v["valor"] for v in tt["valores"]
              if v["variable"] == "abundancia_rel"]
        check("Todo δ13C extraído cae en el rango físico (-75 a 0 ‰)",
              all(-75 <= v < 0 for v in d13), f"{[v for v in d13 if not -75 <= v < 0][:4]}")
        check("Toda abundancia extraída está entre 0 y 100 %",
              all(0 < v <= 100 for v in ab), f"{[v for v in ab if not 0 < v <= 100][:4]}")
        LIM = {"shannon_H": (0, 6), "equidad_J": (0, 1), "simpson": (0, 1),
               "fisher_alpha": (0, 120)}
        malos = [x for x in tt["indices"]
                 if not LIM[x["indice"]][0] <= x["valor"] <= LIM[x["indice"]][1]]
        check("Todo índice de diversidad cae en su rango", not malos,
              f"{[(x['indice'], x['valor']) for x in malos][:4]}")
        cuant = DERIV / "cuantitativos.json"
        if cuant.exists():
            check("El agregado público de tablas no lleva texto literal",
                  "evidencia" not in cuant.read_text(encoding="utf-8"), "")

    # La evidencia de dominancia es texto LITERAL de artículos con derechos de
    # autor. Vive en data/private/ y no puede salir de ahí: si se copiara a un
    # dataset público, el dashboard estaría redistribuyendo el texto ajeno.
    fuga_ev = [f.name for f in DERIV.glob("*.json")
               if "evidencia_dominancia" in f.read_text(encoding="utf-8")]
    check("La evidencia literal de los artículos no sale a data/derived",
          not fuga_ev, f"{fuga_ev}")

    vign = ROOT / ".vercelignore"
    check("Existe .vercelignore (el CLI de Vercel ignora .gitignore)",
          vign.exists(), "sin él, `vercel deploy` sube la carpeta entera")
    if vign.exists():
        reglas = {l.strip().rstrip("/") for l in vign.read_text(encoding="utf-8")
                  .splitlines() if l.strip() and not l.startswith("#")}
        faltan = [d for d in ("Data_nosubiralrepo", "data/private", "*.xlsx")
                  if d not in reglas]
        check("`.vercelignore` cubre los datos primarios inéditos",
              not faltan, f"faltan {faltan}")

    # Un color en hex fijo dentro de un componente NO cambia con el tema. Si se
    # pinta sobre un fondo que sí cambia, el contraste se invierte al pasar a
    # oscuro: así la matriz llegó a mostrar la cifra a contraste 1,00 —fondo y
    # tinta idénticos, número invisible— durante toda la vida del proyecto.
    # Los colores viven en globals.css, que sí tiene los tres bloques de tema.
    comp = ROOT / "src" / "components"
    if comp.is_dir():
        import re as _re
        culpables = []
        for f in sorted(comp.glob("*.tsx")):
            for i, l in enumerate(f.read_text(encoding="utf-8").splitlines(), 1):
                if _re.search(r'"#[0-9a-fA-F]{3,8}"', l):
                    culpables.append(f"{f.name}:{i}")
        check("Ningún componente fija un color en hex (no seguiría al tema)",
              not culpables, f"{culpables}")

        # El separador decimal NO es el mismo en los dos idiomas, y el proyecto
        # llegó a tener tres criterios a la vez: helper propio, `.replace` a
        # mano —que dejaba coma también en inglés— y `.toFixed()` crudo —que
        # dejaba punto también en español—. En producción se leía «H' 3.43» en
        # la versión española. Ahora todo pasa por `useNum()` de src/lib/ui.tsx.
        # Se permite dentro de un par tx({es, en}), donde cada rama ya elige.
        clavados = []
        for f in sorted(comp.glob("*.tsx")):
            for i, l in enumerate(f.read_text(encoding="utf-8").splitlines(), 1):
                if 'replace(".", ",")' in l or 'toLocaleString("es")' in l:
                    clavados.append(f"{f.name}:{i}")
        check("Ningún componente clava el separador decimal o de millar",
              not clavados, f"{clavados}")

    # Crédito académico. «Barragán-Jacksson» es un apellido COMPUESTO y partirlo
    # atribuye el trabajo a otra persona. El dashboard citaba «Barragán y
    # Bernal» en la sección que resume entero su artículo de 2024, y el estudio
    # ni siquiera aparecía en la lista de referencias. Verificado contra la
    # portada de Puerres et al. (2022): Camila María Barragán-Jacksson.
    partido = []
    for f in sorted(ROOT.glob("*.md")) + sorted((ROOT / "src").rglob("*.tsx"))             + sorted((ROOT / "pipeline").glob("*.py")):
        # Este archivo se excluye: para explicar la regla tiene que citar
        # justamente la forma que prohíbe.
        if f.name == "99_auditoria.py":
            continue
        for i, l in enumerate(f.read_text(encoding="utf-8").splitlines(), 1):
            for m in re.finditer(r"Barrag[áa]n", l):
                if not l[m.end():].startswith("-Jacksson"):
                    partido.append(f"{f.name}:{i}")
    check("El apellido Barragán-Jacksson nunca aparece partido",
          not partido, f"{partido}")

    # Y su estudio de 2024 tiene que estar en la lista de referencias, que es
    # donde un lector busca las fuentes. Se mantiene fuera de estudios.json a
    # propósito —no debe llenar la celda tropical somera—, pero eso es no
    # analizarlo, no dejar de citarlo.
    refs = (ROOT / "src" / "components" / "Referencias.tsx").read_text(encoding="utf-8")
    check("El estudio de 2024 se cita en la lista de referencias",
          "sinu_2024.json" in refs,
          "Referencias.tsx debe leer sinu_2024.json")

    # Cada localidad caribeña son cifras de OTRO autor, tomadas de la prosa de
    # la tesis. Sin `fuente` y sin la página donde la tesis las cita, el dato
    # queda huérfano y el dashboard se lo estaría apropiando.
    car = load(DERIV / "caribe_referencia.json")
    sin_fuente = [l["id"] for l in car
                  if l["id"] != "msh_bc21" and not l.get("fuente")]
    sin_cita = [l["id"] for l in car
                if l["id"] != "msh_bc21" and not l.get("cita_en_tesis")]
    check("Toda localidad caribeña declara su fuente", not sin_fuente, f"{sin_fuente}")
    check("Y dónde la cita la tesis", not sin_cita, f"{sin_cita}")
    car_tsx = (ROOT / "src" / "components" / "Caribe.tsx").read_text(encoding="utf-8")
    check("La sección 07 muestra esas fuentes",
          "cita_en_tesis" in car_tsx and "s.fuente" in car_tsx,
          "Caribe.tsx debe renderizar fuente y cita_en_tesis")

    # ---------------------------------------------------------------------
    # Los Excel corregidos son un producto más del pipeline, y hasta ahora
    # nadie los verificaba: se comprueban si existen. Que la hoja de
    # asociaciones cuadre con el pipeline es lo que impide que se quede a
    # medias sin que nadie lo note.
    libro = DATA_DIR / "BD FORAMS AMBTE FILTRACION - CORREGIDA.xlsx"
    if libro.exists():
        print("\n[8] EXCEL CORREGIDOS")
        wbc = openpyxl.load_workbook(libro, data_only=True)
        pdfs = PRIV / "taxones_pdf.json"
        if "Asociaciones por estudio" in wbc.sheetnames and pdfs.exists():
            wsa = wbc["Asociaciones por estudio"]
            n_xl = wsa.max_row - 1
            regs = load(pdfs)["registros"]
            check("La hoja de asociaciones conserva todos los pares estudio-taxón",
                  n_xl == len(regs), f"Excel {n_xl} vs pipeline {len(regs)}")
            # La columna se localiza por su NOMBRE. Con el índice fijo, añadir
            # una columna delante dejaba la comprobación contando otra cosa —
            # y así pasó: leía 0 dominancias donde había 148.
            cab = [c.value for c in wsa[1]]
            col = cab.index("¿Dominante declarado?")
            dom_xl = sum(1 for f in wsa.iter_rows(min_row=2, values_only=True)
                         if f[col] == "sí")
            dom_pp = sum(1 for r in regs if r["dominante_declarado"])
            check("Las dominancias declaradas cuadran con el pipeline",
                  dom_xl == dom_pp, f"Excel {dom_xl} vs pipeline {dom_pp}")
        else:
            check("El Excel corregido trae la hoja de asociaciones",
                  False, "falta «Asociaciones por estudio»")

        col = DATA_DIR / "Coleccion MSH-BC-21 - CORREGIDA.xlsx"
        if col.exists():
            wsc = openpyxl.load_workbook(col, data_only=True)["Clasificación corregida"]
            cab = [c.value for c in wsc[1]]
            if "Pi*Ln(Pi)" in cab:
                i = cab.index("Pi*Ln(Pi)")
                suma = -sum(f[i] for f in wsc.iter_rows(min_row=2, values_only=True))
                h = load(DERIV / "msh_bc21.json")["indices"]["shannon_H"]
                check("Pi*Ln(Pi) del Excel reconstruye el Shannon publicado",
                      abs(suma - h) < 1e-3, f"{suma:.4f} vs {h}")

    # ---------------------------------------------------------------------
    print("\n" + "=" * 74)
    print(f"RESULTADO: {N_OK} comprobaciones OK, {len(FALLAS)} fallas")
    if FALLAS:
        print("\nFALLAS:")
        for f in FALLAS:
            print(f"  - {f}")
    print("=" * 74)
    return 1 if FALLAS else 0


if __name__ == "__main__":
    raise SystemExit(main())
