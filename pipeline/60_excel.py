"""
60_excel.py — Escribe las bases de datos corregidas en Excel.

Hasta ahora todas las correcciones vivían en el pipeline y en los JSON. Esto
las devuelve al formato en el que se trabajó la tesis, para poder seguir
usándolas fuera del dashboard.

NO se tocan los archivos originales. Se generan copias nuevas junto a ellos,
con el sufijo «CORREGIDA», y ambas quedan en la carpeta privada porque
contienen datos primarios inéditos.

Cada libro abre con una hoja «Léeme» que explica qué cambió y qué no, para
que el archivo se entienda sin este código delante.
"""

from __future__ import annotations

import json
import math
import os
import sys
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.environ.get("THESIS_DATA_DIR", ROOT / "Data_nosubiralrepo"))
PRIV = ROOT / "data" / "private"
DERIV = ROOT / "data" / "derived"

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

CAB = PatternFill("solid", fgColor="1F2A37")
CAB_F = Font(color="FFFFFF", bold=True, size=10)
TIT = Font(bold=True, size=13)
NOTA = Font(italic=True, size=9, color="666666")


def hoja(wb: Workbook, nombre: str, cabeceras: list[str], filas: list[list],
         anchos: list[int] | None = None) -> None:
    ws = wb.create_sheet(nombre)
    ws.append(cabeceras)
    for c in ws[1]:
        c.fill, c.font = CAB, CAB_F
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for f in filas:
        ws.append(f)
    for i, a in enumerate(anchos or [22] * len(cabeceras), start=1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def leeme(wb: Workbook, titulo: str, parrafos: list[str]) -> None:
    ws = wb.create_sheet("Léeme", 0)
    ws.column_dimensions["A"].width = 112
    ws["A1"] = titulo
    ws["A1"].font = TIT
    r = 3
    for p in parrafos:
        ws.cell(r, 1, p).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 13 * (len(p) // 100 + 1))
        r += 1
    ws.cell(r + 1, 1, f"Generado automáticamente el {date.today():%d/%m/%Y} por "
                      "el pipeline del proyecto.").font = NOTA


def bibliografia() -> Path:
    bib = json.loads((PRIV / "bibliografia_clean.json").read_text(encoding="utf-8"))
    est = {e["titulo"]: e for e in
           json.loads((DERIV / "estudios.json").read_text(encoding="utf-8"))}
    xref = {x["titulo_original"]: x["titulo_limpio"]
            for x in json.loads((PRIV / "estudios_crossref.json").read_text(encoding="utf-8"))}
    corr = json.loads((DERIV / "correcciones.json").read_text(encoding="utf-8"))
    comp = json.loads((DERIV / "taxones_completo.json").read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    leeme(wb, "Base de foraminíferos en ambientes de filtración — versión corregida", [
        "Esta es la base bibliográfica de la tesis con las correcciones aplicadas y "
        "con la información que se le añadió después. El archivo original no se ha "
        "modificado: esto es una copia nueva.",
        "",
        "QUÉ SE CORRIGIÓ",
        "· Nombres taxonómicos resueltos contra WoRMS (World Foraminifera Database). "
        "Se distingue el nombre tal como estaba en la hoja original (columna "
        "«taxón original») del nombre válido hoy (columna «taxón»).",
        "· Se eliminaron 5 especies planctónicas que estaban en una base declarada de "
        "bentónicos, y un marcador de posición sin contenido.",
        "· Se eliminaron 12 filas duplicadas exactas: mismo estudio, taxón, banda "
        "latitudinal, profundidad y microhábitat. Se CONSERVARON las repeticiones en "
        "que cambia la banda, porque ahí el artículo reporta la especie en dos "
        "estratos distintos y son observaciones separadas.",
        "· Se corrigió el tipo de pared de 5 taxones cuyo género tiene otra posición "
        "sistemática (Ammodiscus y Glomospira son aglutinados, no porcelanáceos).",
        "",
        "QUÉ SE AÑADIÓ",
        "· Localidad y coordenadas de cada estudio, que la base original no tenía. Son "
        "la posición representativa de la localidad, no la del testigo.",
        "· Tipo de fluido y expresión geomorfológica (pockmark, volcán de lodo…), en "
        "dos columnas independientes porque son ejes distintos.",
        "· Microhábitat, recuperado de la hoja borrador, donde se había perdido.",
        "· Referencia bibliográfica completa con DOI, resuelta contra CrossRef.",
        "· AphiaID, familia y orden de cada taxón.",
        "· LA ASOCIACIÓN COMPLETA DE CADA ARTÍCULO, en la hoja «Asociaciones por "
        "estudio»: 1.527 pares estudio-taxón leídos del texto completo. La base de la "
        "tesis recogía «las 5 principales especies» de cada filtro, es decir una "
        "muestra de las dominantes; esto es todo lo que cada artículo nombra.",
        "· «Resumen por estudio»: cuántos taxones, especies y géneros aporta cada "
        "artículo, y cuáles declara dominantes.",
        "",
        "CÓMO SE DECIDE QUÉ REPORTA CADA ARTÍCULO",
        "· La bibliografía se descarta antes de contar. La lista de referencias "
        "cita los títulos de otros trabajos, y esos títulos llevan nombres de "
        "especies: contarlos daba por «reportado» un taxón que el artículo sólo "
        "nombraba al citar a un tercero. La columna «menciones en la "
        "bibliografía» deja ver cuántas se descartaron en cada caso.",
        "· Los nombres abreviados se expanden: «U. peregrina» cuenta como "
        "«Uvigerina peregrina». Los artículos escriben el binomio entero la "
        "primera vez y lo abrevian después, sobre todo en tablas. Sólo se "
        "expande cuando inicial y epíteto identifican un único binomio del "
        "propio artículo; si «B. spissa» pudiera ser Bolivina o Bulimina, se "
        "deja como está antes que adivinar.",
        "· «¿Aparece en Resultados?» distingue el taxón que el artículo mide del "
        "que sólo menciona al repasar la literatura. Donde no se pudo localizar "
        "el encabezado de Resultados dice «sin determinar», que no es lo mismo "
        "que «no».",
        "· La dominancia se atribuye por CLÁUSULA, no por frase. Antes bastaba "
        "con que el taxón y la palabra «dominant» cayeran en la misma frase, y "
        "eso marcaba dominante a cualquier taxón nombrado de paso: en «Bolivina "
        "dominated the assemblage, whereas Uvigerina was rare», Uvigerina salía "
        "dominante. Ahora se descartan además las cláusulas negadas («X was not "
        "dominant») y las que atribuyen el hallazgo a otro trabajo («Rathburn et "
        "al. (2000) found X dominant»). La columna «frase que lo afirma» trae la "
        "cláusula literal para que la decisión se pueda auditar una a una.",
        "",
        "CÓMO SE LEEN LAS TRES SEÑALES, QUE NO VALEN LO MISMO",
        "· Presencia: el taxón aparece en el artículo. Señal sólida.",
        "· Menciones: cuántas veces se nombra. Proxy DÉBIL de importancia — un "
        "artículo repite un nombre en la introducción, en la discusión y al citar a "
        "otros. Sirve para ordenar dentro de un artículo, no para comparar entre "
        "artículos de distinta extensión.",
        "· Dominante declarado: el propio texto lo afirma («dominant», «most "
        "abundant»…). Es la única señal que puede llamarse dominancia.",
        "",
        "QUÉ NO CONTIENE — LÍMITES QUE CONVIENE CONOCER",
        "· Abundancias numéricas por especie. La base original registra presencia, no "
        "cantidad. De los 36 artículos legibles, 33 mencionan abundancias relativas, "
        "25 traen valores de δ13C y 9 reportan índices de diversidad: NADA DE ESO SE "
        "EXTRAJO. Vive en tablas que el lector de PDF no interpreta de forma fiable, "
        "y extraerlas mal sería peor que no extraerlas.",
        "· 14 de los 36 artículos NO declaran qué foraminífero domina. No es un "
        "fallo de la extracción: usan «dominant» para la litología («the dominant "
        "lithology is sandy clay»), para procesos geoquímicos («dominated by AOM "
        "and sulfate reduction»), para tapetes bacterianos («dominated by "
        "Beggiatoa») o para gusanos tubícolas («the second most abundant tube "
        "worm»). Para esos estudios la hoja «Asociaciones dominantes» ofrece los "
        "taxones más mencionados en Resultados, y la columna «Origen» avisa de "
        "que es un indicio DERIVADO, no una afirmación del artículo.",
        "· Las abundancias del CUERPO DEL TEXTO no se extrajeron como cifras, y "
        "es deliberado: en varios artículos el símbolo ‰ de los valores de δ13C "
        "sale del PDF convertido en «%», de modo que «1,26 ± 0,15 %» es un valor "
        "isotópico y no una abundancia. Lo que sí se extrajo son las TABLAS, con "
        "otra herramienta y con validación de rango (hojas «δ13C por taxón», "
        "«Abundancias de los artículos» e «Índices de diversidad»).",
        "· La cobertura de esas tres hojas es PARCIAL y desigual, porque muchas "
        "tablas son imágenes o tienen la cabecera de columna ilegible. Cada valor "
        "trae la página y la fila literal del artículo para poder comprobarlo.",
        "· De las abundancias no se dice a qué estación corresponde cada cifra: "
        "los encabezados de columna no se leen de forma fiable. Una fila trae "
        "todos los valores de ese taxón en esa tabla.",
        "· De los isótopos sólo se recogen los valores NEGATIVOS de δ13C. Es la "
        "firma del carbono derivado del metano —lo que interesa aquí— y es lo "
        "único que permite distinguir un δ13C de un δ18O cuando la cabecera no se "
        "puede leer. Los positivos quedan sin clasificar.",
        "· 7 de los 40 estudios no tienen morfología asignada. En varios el texto sí "
        "menciona pockmarks o volcanes de lodo, pero la mención está en una tabla "
        "comparativa de otras localidades o en la bibliografía, no en la descripción "
        "del sitio muestreado. Asignarla sería inventar.",
        "",
        "La hoja «Correcciones» lista una a una todas las diferencias respecto del "
        "archivo original, con su motivo y su fuente.",
    ])

    # δ13C leído de las tablas, indexado por (estudio, taxón). Va a la BASE
    # PRINCIPAL y no a una hoja aparte porque comparte su mismo grano —una
    # medida por taxón dentro de un estudio— y el 73 % de los valores cae sobre
    # registros que la base ya tiene. Las abundancias, en cambio, sólo solapan
    # un 17 %: la base recoge «las 5 principales especies» por filtro y las
    # tablas listan la asociación entera, así que van a la vista ampliada.
    cuant = json.loads((DERIV / "cuantitativos.json").read_text(encoding="utf-8")) \
        if (DERIV / "cuantitativos.json").exists() else {"d13C": [], "abundancia_rel": []}
    d13 = {(x["estudio_id"], x["taxon"]): x for x in cuant["d13C"]}
    abund = {(x["estudio_id"], x["taxon"]): x for x in cuant["abundancia_rel"]}

    filas: list[list] = []
    for r in bib:
        e = est.get(xref.get(r["estudio"], ""), {})
        iso = d13.get((e.get("id"), r["taxon"]))
        filas.append([
            e.get("id"), (e.get("autores") or [None])[0], e.get("anio"),
            e.get("titulo") or r["estudio"], e.get("doi"),
            e.get("localidad"), e.get("region"), e.get("lat"), e.get("lon"),
            e.get("prof_m"), e.get("tipo_filtracion"), e.get("morfologia_label"),
            r["taxon"], r["taxon_original"], r["aphia_id"], r["rango"],
            r["genero_label"], r["familia"], r["orden"],
            r["pared"], r["subtipo"], r["lat_banda"], r["prof_banda"],
            r.get("microhabitat_label"),
            "sí" if r.get("recuperado") else "no",
            "derivada" if r.get("pared_derivada") else "original",
            "sí" if r["verificado_worms"] else "no",
            iso["min"] if iso else None,
            iso["max"] if iso else None,
            iso["n"] if iso else None,
        ])
    hoja(wb, "Base corregida", [
        "ID estudio", "Primer autor", "Año", "Título", "DOI",
        "Localidad", "Región", "Latitud", "Longitud", "Profundidad (m)",
        "Tipo de fluido", "Morfología",
        "Taxón (nombre válido)", "Taxón original", "AphiaID", "Rango",
        "Género", "Familia", "Orden",
        "Tipo de pared", "Subtipo", "Banda latitudinal", "Banda de profundidad",
        "Microhábitat", "Estudio reincorporado", "Origen del tipo de pared",
        "Verificado en WoRMS",
        "δ13C mín (‰)", "δ13C máx (‰)", "δ13C nº valores",
    ], filas, [10, 18, 7, 46, 26, 34, 16, 10, 10, 12, 16, 22, 30, 30, 10, 10, 18, 20, 16, 14, 15, 16, 18, 20, 18, 18, 16, 13, 13, 13])

    hoja(wb, "Estudios", [
        "ID", "Autores", "Año", "Título", "Revista", "DOI", "Registros",
        "Localidad", "Región", "Latitud", "Longitud", "Profundidad (m)",
        "Tipo de fluido", "Morfología", "Confianza de la localidad",
        "Fuente de la localidad", "¿Es filtración?", "Reincorporado",
    ], [[
        e["id"], ", ".join(e.get("autores") or []), e.get("anio"), e["titulo"],
        e.get("revista"), e.get("doi"), e["n_registros"], e.get("localidad"),
        e.get("region"), e.get("lat"), e.get("lon"), e.get("prof_m"),
        e.get("tipo_filtracion"), e.get("morfologia_label"), e.get("confianza"),
        e.get("fuente"), "sí" if e["es_filtracion"] else "no",
        "sí" if e.get("recuperado") else "no",
    ] for e in sorted(est.values(), key=lambda x: -x["n_registros"])],
        [8, 40, 7, 50, 30, 26, 10, 34, 16, 10, 10, 12, 16, 22, 12, 60, 14, 13])

    hoja(wb, "Taxones (base)", [
        "Taxón", "Rango", "Género", "Familia", "AphiaID", "Registros",
        "Estudios", "Bandas latitudinales", "Bandas de profundidad",
        "Microhábitats", "Tipo de pared", "Subtipo", "Verificado en WoRMS",
    ], [[
        t["taxon"], t["rango"], t["genero"], t["familia"], t["aphia_id"],
        t["registros"], t["n_estudios"], ", ".join(t["lats"]),
        ", ".join(t["profs"]), ", ".join(t["microhabitats"]),
        t["pared"], t["subtipo"], "sí" if t["verificado_worms"] else "no",
    ] for t in json.loads((DERIV / "taxones_global.json").read_text(encoding="utf-8"))],
        [32, 10, 20, 22, 10, 10, 10, 22, 24, 24, 14, 16, 16])

    # Qué estudio reporta cada taxón. El recuento solo no permite rastrear la
    # afirmación hasta su artículo, y esta base existe para poder hacerlo.
    reg = json.loads(
        (PRIV / "taxones_pdf.json").read_text(encoding="utf-8"))["registros"]
    quien: dict[str, list[str]] = {}
    for r in reg:
        quien.setdefault(r["taxon"], []).append(r["estudio_id"])
    dom: dict[str, list[str]] = {}
    for r in reg:
        if r["dominante_declarado"]:
            dom.setdefault(r["taxon"], []).append(r["estudio_id"])

    hoja(wb, "Taxones en los artículos", [
        "Taxón", "Rango", "Género", "Familia", "Orden", "AphiaID",
        "Nº de estudios que lo reportan", "Reportado en (IDs)",
        "Estudios donde aparece en Resultados",
        "Declarado dominante en", "Dominante en (IDs)",
        "¿Estaba en la base?", "¿Está en MSH-BC-21?",
    ], [[
        t["taxon"], t["rango"], t["genero"], t["familia"], t["orden"],
        t["aphia_id"], t["n_estudios"],
        ", ".join(sorted(set(quien.get(t["taxon"], [])))),
        t.get("n_estudios_resultados"),
        t["n_estudios_dominante"],
        ", ".join(sorted(set(dom.get(t["taxon"], [])))),
        "sí" if t["en_base_tesis"] else "no",
        "sí" if t["en_msh_bc21"] else "no",
    ] for t in comp["taxones"]],
        [32, 10, 20, 22, 16, 10, 16, 30, 18, 16, 26, 16, 16])

    # LA ASOCIACIÓN COMPLETA, estudio por estudio. Es el dato que la tesis no
    # tenía: su base recogía «las 5 principales especies» de cada filtro, de
    # modo que era una muestra de las dominantes, no la asociación entera.
    est_por_id = {e["id"]: e for e in
                  json.loads((DERIV / "estudios.json").read_text(encoding="utf-8"))}
    en_base = {t["taxon"] for t in
               json.loads((DERIV / "taxones_global.json").read_text(encoding="utf-8"))}
    filas_asoc = []
    for r in sorted(reg, key=lambda x: (x["estudio_id"], -x["menciones"], x["taxon"])):
        e = est_por_id.get(r["estudio_id"], {})
        filas_asoc.append([
            r["estudio_id"], (e.get("autores") or [""])[0], e.get("anio"),
            e.get("localidad"), e.get("tipo_filtracion"), e.get("morfologia_label"),
            r["taxon"], r["taxon_texto"], r["rango"], r["genero"], r["familia"],
            r["orden"], r["aphia_id"], r["menciones"],
            r.get("menciones_en_referencias", 0),
            {True: "sí", False: "no", None: "sin determinar"}[
                r.get("en_resultados")],
            "sí" if r["dominante_declarado"] else "no",
            r.get("evidencia_dominancia") or "",
            "sí" if r["taxon"] in en_base else "no",
            (ab := abund.get((r["estudio_id"], r["taxon"]))) and ab["min"],
            ab["max"] if ab else None,
            ab["n"] if ab else None,
        ])
    hoja(wb, "Asociaciones por estudio", [
        "ID estudio", "Primer autor", "Año", "Localidad", "Tipo de fluido",
        "Morfología", "Taxón (nombre válido)", "Nombre leído en el texto",
        "Rango", "Género", "Familia", "Orden", "AphiaID", "Menciones",
        "Menciones en la bibliografía", "¿Aparece en Resultados?",
        "¿Dominante declarado?", "Frase que lo afirma",
        "¿Estaba en la base de la tesis?",
        "Abund. mín (%)", "Abund. máx (%)", "Abund. nº valores",
    ], filas_asoc,
        [11, 18, 7, 34, 14, 22, 32, 32, 10, 20, 22, 16, 10, 11, 14, 16, 16,
         80, 20, 14, 14, 15])

    # Una fila por estudio con SU asociación dominante. Es la pregunta directa
    # —«qué domina en cada localidad»— y hasta ahora había que reconstruirla
    # filtrando 1.527 filas a mano.
    # Todos los estudios aparecen, pero la columna «Origen» dice si la
    # dominancia la AFIRMA el artículo o la derivamos nosotros. Sin esa
    # distinción, la hoja daría por declarado lo que sólo es un indicio.
    hoja(wb, "Asociaciones dominantes", [
        "ID estudio", "Autores", "Año", "Localidad", "Tipo de fluido",
        "Morfología", "Origen", "Nº de taxones",
        "Asociación dominante (o taxones destacados)",
    ], [[
        p["estudio_id"], p["autores"], p["anio"], p["localidad"],
        p["tipo_filtracion"], p["morfologia"],
        "declarada por el artículo" if p.get("origen_dominancia") == "declarada"
        else "DERIVADA: más mencionados (el artículo no la declara)",
        len(p.get("asociacion_dominante")
            or p.get("destacados_derivados") or []),
        ", ".join(p.get("asociacion_dominante")
                  or p.get("destacados_derivados") or []) or "—",
    ] for p in comp["por_estudio"]],
        [11, 24, 7, 34, 14, 22, 46, 12, 96])

    hoja(wb, "Resumen por estudio", [
        "ID", "Autores", "Año", "Título", "Localidad", "Tipo de fluido",
        "Morfología", "Nº de taxones", "Nº de especies", "Nº de géneros",
        "De ellos, en Resultados", "Registros en la base de la tesis",
        "Dominantes declarados",
    ], [[
        p["estudio_id"], p["autores"], p["anio"], p["titulo"], p["localidad"],
        p["tipo_filtracion"], p["morfologia"], p["n_taxones"], p["n_especies"],
        p["n_generos"], p.get("n_en_resultados"), p["n_en_base_tesis"],
        ", ".join(p["dominantes_declarados"]) or "—",
    ] for p in comp["por_estudio"]],
        [8, 24, 7, 60, 34, 14, 22, 12, 12, 12, 14, 16, 46])

    # --- lo extraído de las TABLAS (45_tablas_pdf.py) --------------------
    tab = PRIV / "tablas_pdf.json"
    if tab.exists():
        tt = json.loads(tab.read_text(encoding="utf-8"))
        est_id = {e["id"]: e for e in
                  json.loads((DERIV / "estudios.json").read_text(encoding="utf-8"))}

        def fila_val(v):
            e = est_id.get(v["estudio_id"], {})
            return [v["estudio_id"], (e.get("autores") or [""])[0], e.get("anio"),
                    e.get("localidad"), v["taxon"], v["taxon_texto"],
                    v["valor"], v["unidad"], v["pagina"], v["evidencia"]]

        cab_val = ["ID estudio", "Primer autor", "Año", "Localidad",
                   "Taxón (nombre válido)", "Nombre leído", "Valor", "Unidad",
                   "Página", "Fila del artículo"]
        anchos = [11, 18, 7, 30, 30, 28, 10, 10, 8, 84]

        hoja(wb, "δ13C por taxón",
             cab_val,
             [fila_val(v) for v in tt["valores"] if v["variable"] == "d13C"],
             anchos)

        hoja(wb, "Abundancias de los artículos",
             cab_val,
             [fila_val(v) for v in tt["valores"]
              if v["variable"] == "abundancia_rel"],
             anchos)

        NOM = {"shannon_H": "Shannon-Wiener (H')", "equidad_J": "Equidad (J')",
               "simpson": "Simpson", "fisher_alpha": "Fisher alfa"}
        hoja(wb, "Índices de diversidad", [
            "ID estudio", "Primer autor", "Año", "Localidad", "Índice",
            "Valor", "Página", "Frase del artículo",
        ], [[
            x["estudio_id"],
            (est_id.get(x["estudio_id"], {}).get("autores") or [""])[0],
            est_id.get(x["estudio_id"], {}).get("anio"),
            est_id.get(x["estudio_id"], {}).get("localidad"),
            NOM.get(x["indice"], x["indice"]), x["valor"], x["pagina"],
            x["evidencia"],
        ] for x in tt["indices"]], [11, 18, 7, 30, 22, 10, 8, 92])

    hoja(wb, "Correcciones", [
        "Tipo", "Desde", "Hacia", "Motivo", "Fuente", "Impacto",
        "Confianza", "Registros afectados",
    ], [[c["tipo"], c["desde"], c["hacia"], c["motivo"], c["fuente"],
         c["impacto"], c["confianza"], c["registros_afectados"]]
        for c in corr["correcciones"]], [22, 40, 32, 76, 30, 44, 12, 12])

    salida = DATA_DIR / "BD FORAMS AMBTE FILTRACION - CORREGIDA.xlsx"
    wb.save(salida)
    return salida


def coleccion() -> Path:
    msh = json.loads((DERIV / "msh_bc21.json").read_text(encoding="utf-8"))
    raw = json.loads((PRIV / "msh_raw.json").read_text(encoding="utf-8"))
    corr = json.loads((DERIV / "correcciones.json").read_text(encoding="utf-8"))
    sinu = json.loads((DERIV / "sinu_2024.json").read_text(encoding="utf-8"))
    compartidos = set(sinu["taxones_compartidos_con_msh"])

    wb = Workbook()
    wb.remove(wb.active)

    leeme(wb, "Colección MSH-BC-21 — versión corregida", [
        "Conteo de foraminíferos de los dos centímetros superficiales del testigo "
        "MSH-BC-21, con las correcciones aplicadas. El archivo original no se ha "
        "modificado: esto es una copia nueva.",
        "",
        "QUÉ SE CORRIGIÓ",
        "· Ammodiscus sp. estaba clasificado como calcáreo porcelanáceo. Es un género "
        "AGLUTINADO. Es la corrección de mayor impacto: la proporción de formas "
        "calcáreas frente a aglutinadas pasa de 88,8 / 11,2 a 87,0 / 13,0.",
        "· Spirillina sp. estaba como porcelanáceo. Su pared es calcárea "
        "monocristalina (Spirillinida), ni hialina ni porcelanácea. No altera la "
        "razón calcáreos/aglutinados, sólo el desglose de subtipos.",
        "· Los nombres se resolvieron contra WoRMS y se añadió el AphiaID.",
        "· Los índices se recalcularon desde los conteos. El Shannon coincide con el "
        "de la hoja original hasta el cuarto decimal (3,4325 frente a 3,4327).",
        "",
        "SOBRE EL TEXTO DE LA TESIS",
        "· La tesis dice «cerca de un 80%» de predominancia de calcáreos sobre "
        "aglutinados. El valor que arrojan los datos es 87,0%. Aquí se publica el "
        "valor calculado.",
        "· La suma de abundancias relativas de la hoja original daba 1,0001 por "
        "redondeo acumulado; aquí se recalcula desde los conteos.",
        "· El total de la hoja «Clasificación» (1214,125) y la suma de la hoja "
        "«Gráficas» (1213,9375) diferían en 0,1875. Se conserva el primero, que es "
        "el que cuadra con los conteos.",
        "",
        "QUÉ NO SE PUDO RECONSTRUIR",
        "· Los conteos son fraccionarios porque proceden de submuestreo por alícuotas, "
        "pero el factor de reparto no está documentado en ninguna hoja del archivo "
        "original. La cadena desde el conteo crudo hasta el total ponderado no es "
        "reproducible con lo que hay.",
        "· La hoja original no desglosa por centímetro ni por fracción de tamaño: los "
        "conteos por especie están agregados. Esa resolución se perdió.",
        "",
        "La columna «También en el Sinú (2024)» marca las especies que Barragán y "
        "Bernal reportan en el mismo campo de filtración, que es el contraste más "
        "directo disponible para esta muestra.",
    ])

    # Se devuelve Pi*Ln(Pi) por especie, que la hoja original sí traía: es el
    # sumando de Shannon, y sin él no se puede rehacer el índice a mano.
    hoja(wb, "Clasificación corregida", [
        "Taxón (nombre válido)", "Taxón original", "AphiaID", "Género", "Familia",
        "Tipo de pared", "Subtipo", "Conteo ponderado",
        "Abundancia relativa (%)", "Pi", "Pi*Ln(Pi)", "También en el Sinú (2024)",
    ], [[
        e["taxon"], e["taxon_original"], e["aphia_id"], e["genero"], e["familia"],
        e["pared"], e["subtipo"], e["conteo"], e["abundancia_rel"],
        round(pi := e["conteo"] / msh["total_ponderado"], 6),
        round(pi * math.log(pi), 6),
        "sí" if e["taxon"] in compartidos else "",
    ] for e in msh["especies"]],
        [32, 32, 10, 20, 24, 14, 16, 16, 18, 12, 14, 20])

    ind = msh["indices"]
    hoja(wb, "Índices y composición", ["Indicador", "Valor", "Nota"], [
        ["Riqueza específica (S)", ind["riqueza_S"], "número de taxones"],
        ["Shannon-Wiener (H')", ind["shannon_H"],
         "recalculado; la hoja original daba 3,4327"],
        ["Equidad de Pielou (J')", ind["equidad_J"], "H' / ln(S); no estaba en el original"],
        ["Dominancia de Simpson (D)", ind["simpson_D"], "no estaba en el original"],
        ["Abundancia en las 5 principales (%)", ind["dominancia_top5"], ""],
        ["", "", ""],
        ["Calcáreos (%)", msh["pared"]["Calcareo"], "era 88,8 antes de corregir Ammodiscus"],
        ["Aglutinados (%)", msh["pared"]["Aglutinado"], "era 11,2 antes de corregir Ammodiscus"],
        ["", "", ""],
        *[[f"  {k} (%)", v, ""] for k, v in msh["subtipo"].items()],
        ["", "", ""],
        ["Especies calcáreas", msh["pared_n_especies"].get("Calcareo"), ""],
        ["Especies aglutinadas", msh["pared_n_especies"].get("Aglutinado"), ""],
    ], [40, 16, 56])

    fr = ["125", "250", "500", "TOTAL"]
    filas: list[list] = []
    ab = {}
    for row in raw["abundancias"]:
        ab.setdefault(row["muestra"], {})[row["variable"]] = row["valores"]
    for m, vs in ab.items():
        for var, vals in vs.items():
            filas.append([m, var] + [vals.get(f) for f in fr])
    hoja(wb, "Abundancias", ["Muestra", "Variable", "125 µm", "250 µm", "500 µm",
                             "TOTAL"], filas, [22, 30, 14, 14, 14, 14])

    hoja(wb, "Correcciones", [
        "Tipo", "Desde", "Hacia", "Motivo", "Fuente", "Impacto",
    ], [[c["tipo"], c["desde"], c["hacia"], c["motivo"], c["fuente"], c["impacto"]]
        for c in corr["correcciones"]
        if c["archivo"] in ("B", "manuscrito")
        or "Ammodiscus" in c["desde"] or "Spirillina" in c["desde"]],
        [22, 40, 32, 80, 30, 46])

    salida = DATA_DIR / "Coleccion MSH-BC-21 - CORREGIDA.xlsx"
    wb.save(salida)
    return salida


def main() -> int:
    a = bibliografia()
    b = coleccion()
    print("EXCEL CORREGIDOS GENERADOS\n")
    for p in (a, b):
        print(f"  {p.name}")
        print(f"    {p.stat().st_size / 1024:.0f} KB")
    print("\nLos archivos originales no se han modificado.")
    print("Ambos quedan en la carpeta privada: contienen datos primarios inéditos.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
